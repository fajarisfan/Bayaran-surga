import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook

st.set_page_config(page_title="Krakatau Baja Master Rekap", page_icon="🏗️")

st.title("🏗️ Master Data Karyawan (Untuk VLOOKUP)")
st.markdown("---")
st.write("Upload semua file BA Bulanan, script ini akan menggabungkan **SEMUA NAMA** ke dalam satu file Excel.")

days_input = st.number_input("Masukkan Jumlah Hari Kerja Efektif per Bulan", min_value=1, value=22)

uploaded_files = st.file_uploader("Pilih file Excel BA", accept_multiple_files=True, type=['xlsx'])

if uploaded_files:
    all_data_frames = []

    for file in uploaded_files:
        all_sheets = pd.read_excel(file, sheet_name=None, header=None)

        for sheet_name, df in all_sheets.items():
            mask = df.astype(str).apply(
                lambda x: x.str.contains('NAMA', case=False, na=False)
            ).any(axis=1)

            if mask.any():  # ✅ FIX: indentasi diperbaiki (sebelumnya ada spasi lebih)
                header_idx = df[mask].index[0]

                df_clean = pd.read_excel(file, sheet_name=sheet_name, skiprows=header_idx)

                df_clean = df_clean.loc[:, ~df_clean.columns.str.contains('^Unnamed')]

                nama_cols = [c for c in df_clean.columns if 'NAMA' in str(c).upper()]

                if nama_cols:
                    col_key = nama_cols[0]
                    df_clean = df_clean.dropna(subset=[col_key])

                    cols_to_keep = ['NO', 'NIK', 'NAMA', 'JABATAN', 'PERUSAHAAN']
                    existing_cols = [c for c in cols_to_keep if c in df_clean.columns]
                    df_clean = df_clean[existing_cols].copy()

                    df_clean['Sumber_File'] = file.name.split('.')[0]
                    df_clean['Mandays_Bulan_Ini'] = days_input
                    all_data_frames.append(df_clean)

    if all_data_frames:
        df_master = pd.concat(all_data_frames, ignore_index=True)

        st.subheader(f"✅ Berhasil Mengumpulkan {len(df_master)} Baris Data")
        st.dataframe(df_master.head(20))

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_master.to_excel(writer, index=False, sheet_name='Master_Data_Zizah')

        # ✅ Format Excel supaya rapi
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(buffer)
        ws = wb['Master_Data_Zizah']

        HEADER_BG = "1F3864"
        THIN = Side(style='thin', color="AAAAAA")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        col_widths = {
            'NO': 6, 'NIK': 13, 'NAMA': 30, 'JABATAN': 18,
            'PERUSAHAAN': 22, 'Sumber_File': 22, 'Mandays_Bulan_Ini': 20
        }
        for i, col_name in enumerate(df_master.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 15)

        # Insert 3 title rows
        ws.insert_rows(1, 3)

        ws.merge_cells('A1:G1')
        ws['A1'].value = 'PT KRAKATAU BAJA'
        ws['A1'].font = Font(name='Arial', bold=True, size=14, color="2E75B6")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:G2')
        ws['A2'].value = 'MASTER DATA KARYAWAN — UNTUK VLOOKUP'
        ws['A2'].font = Font(name='Arial', bold=True, size=12, color="1F3864")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        # Header row (now row 4)
        nice_headers = {
            'NO': 'No.', 'NIK': 'NIK', 'NAMA': 'Nama Karyawan',
            'JABATAN': 'Jabatan', 'PERUSAHAAN': 'Perusahaan',
            'Sumber_File': 'Sumber File', 'Mandays_Bulan_Ini': 'Mandays Bulan Ini'
        }
        for cell in ws[4]:
            cell.value = nice_headers.get(cell.value, cell.value)
            cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        ws.row_dimensions[4].height = 28

        # Data rows
        for r_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
            bg = "DCE6F1" if r_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center')
                if cell.column in [1, 2, 7]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r_idx + 4].height = 20

        ws.freeze_panes = 'A5'

        buffer = io.BytesIO()
        wb.save(buffer)

        st.download_button(
            label="📥 Download Master Data Lengkap (Excel)",
            data=buffer.getvalue(),
            file_name='MASTER_DATA_VLOOKUP_ZIZAH.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        st.success("Download file di atas, isinya sudah daftar nama lengkap untuk bahan VLOOKUP. 🥂")
    else:
        st.error("⚠️ Tidak ditemukan kolom 'NAMA'. Pastikan header di Excel tulisannya 'NAMA'.")

# --- ZONA TESTING ---
st.markdown("---")
st.subheader("🛠️ Zona Testing (Gunakan ini kalau mau tes hasil)")
st.info("Download file di bawah, lalu upload kembali ke atas untuk tes.")

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _format_ba(ws, judul_rows, karyawan, data_start_row):
    """Helper: format sheet BA dengan styling rapi."""
    HEADER_BG  = "1F3864"
    THIN       = Side(style='thin', color="AAAAAA")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    COL_WIDTHS = [6, 13, 30, 18, 24]   # NO, NIK, NAMA, JABATAN, PERUSAHAAN

    # Judul atas (merge A:E)
    for r, (val, bold, size) in enumerate(judul_rows, 1):
        ws.merge_cells(f'A{r}:E{r}')
        c = ws.cell(r, 1, val)
        c.font = Font(name='Arial', bold=bold, size=size, color="1F3864")
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 20

    # Baris kosong sebelum header tabel
    ws.row_dimensions[data_start_row - 1].height = 6

    # Header tabel
    for col, h in enumerate(['NO', 'NIK', 'NAMA', 'JABATAN', 'PERUSAHAAN'], 1):
        c = ws.cell(data_start_row, col, h)
        c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[data_start_row].height = 26

    # Data karyawan
    for i, row in enumerate(karyawan, data_start_row + 1):
        bg = "DCE6F1" if (i - data_start_row) % 2 == 0 else "FFFFFF"
        for j, val in enumerate(row, 1):
            c = ws.cell(i, j, val)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = BORDER
            c.alignment = Alignment(
                horizontal='center' if j in [1, 2] else 'left',
                vertical='center'
            )
        ws.row_dimensions[i].height = 20

    # Lebar kolom
    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f'A{data_start_row + 1}'


def buat_file_test_maret():
    wb = Workbook()
    ws = wb.active
    ws.title = "BA_Maret"
    judul = [
        ('PT KRAKATAU BAJA',                    True,  13),
        ('BERITA ACARA KEHADIRAN TENAGA KERJA',  True,  11),
        ('BULAN: MARET 2026',                    False, 10),
        ('LOKASI KERJA: CILEGON',                False, 10),
        ('',                                     False, 10),
    ]
    karyawan = [
        (1, 'A001', 'Zizah Nur Aini',  'Admin',      'PT Maju Jaya'),
        (2, 'A002', 'Budi Santoso',    'Teknisi',    'PT Maju Jaya'),
        (3, 'A003', 'Ani Rahayu',      'Logistik',   'PT Sejahtera'),
        (4, 'A004', 'Candra Wijaya',   'Supervisor', 'PT Sejahtera'),
        (5, 'A005', 'Dewi Lestari',    'Operator',   'PT Karya Utama'),
    ]
    _format_ba(ws, judul, karyawan, data_start_row=6)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def buat_file_test_april():
    wb = Workbook()
    ws = wb.active
    ws.title = "BA_April"
    judul = [
        ('PT KRAKATAU BAJA',                   True,  13),
        ('BERITA ACARA KEHADIRAN TENAGA KERJA', True,  11),
        ('BULAN: APRIL 2026',                   False, 10),
        ('KONTRAKTOR: PT MAJU BERSAMA  |  LOKASI: CILEGON', False, 10),
    ]
    karyawan = [
        (1, 'B001', 'Dedi Kurniawan',    'Driver',   'PT Logindo'),
        (2, 'B002', 'Eka Putri Mandiri', 'Security', 'PT Logindo'),
        (3, 'B003', 'Faisal Rahman',     'Helper',   'PT Konstruksi'),
        (4, 'B004', 'Gina Marlina',      'Operator', 'PT Konstruksi'),
    ]
    _format_ba(ws, judul, karyawan, data_start_row=5)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

col1, col2 = st.columns(2)

with col1:
    st.markdown("**File Test A — Maret** (5 karyawan)")
    st.download_button(
        label="📥 Download File Test Maret",
        data=buat_file_test_maret(),
        file_name="Test_BA_Maret.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with col2:
    st.markdown("**File Test B — April** (4 karyawan)")
    st.download_button(
        label="📥 Download File Test April",
        data=buat_file_test_april(),
        file_name="Test_BA_April.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.caption("Setelah download kedua file, upload ke uploader di atas untuk melihat hasil rekap.")
