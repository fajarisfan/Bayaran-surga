import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook

st.set_page_config(page_title="Krakatau Baja Master Rekap", page_icon="🏗️")

st.title("🏗️ Master Data Karyawan (Untuk VLOOKUP)")
st.markdown("---")
st.write("Upload semua file BA Bulanan, script ini akan menggabungkan **SEMUA NAMA** ke dalam satu file Excel.")

days_input = st.number_input("Masukkan Jumlah Hari Kerja Efektif per Bulan", min_value=1, value=22)

uploaded_files = st.file_uploader("Pilih file Excel BA", accept_multiple_files=True, type=['xlsx'])

if uploaded_files:
    all_data_frames = []

    for file in uploaded_files:
        # ✅ KUNCI FIX: header=None supaya pandas tidak salah baca struktur file
        all_sheets = pd.read_excel(file, sheet_name=None, header=None)

        for sheet_name, df in all_sheets.items():
            # Cari baris mana yang mengandung kata "NAMA"
            mask = df.astype(str).apply(
                lambda x: x.str.contains('NAMA', case=False, na=False)
            ).any(axis=1)

           if mask.any():
                header_idx = df[mask].index[0]

                # Baca ulang dengan skiprows yang benar
                df_clean = pd.read_excel(file, sheet_name=sheet_name, skiprows=header_idx)

                # ✅ FIX: Buang kolom yang namanya "Unnamed" (kolom sampah)
                df_clean = df_clean.loc[:, ~df_clean.columns.str.contains('^Unnamed')]

                # Identifikasi kolom NAMA secara dinamis
                nama_cols = [c for c in df_clean.columns if 'NAMA' in str(c).upper()]

                if nama_cols:
                    col_key = nama_cols[0]
                    # Bersihkan baris yang namanya benar-benar kosong
                    df_clean = df_clean.dropna(subset=[col_key])
                    
                    # ✅ FIX: Pastikan hanya mengambil kolom yang relevan saja (opsional)
                    # Biar nggak banyak kolom sampah dari file aslinya
                    cols_to_keep = ['NO', 'NIK', 'NAMA', 'JABATAN', 'PERUSAHAAN']
                    existing_cols = [c for c in cols_to_keep if c in df_clean.columns]
                    df_clean = df_clean[existing_cols].copy()

                    df_clean['Sumber_File'] = file.name.split('.')[0]
                    df_clean['Mandays_Bulan_Ini'] = days_input
                    all_data_frames.append(df_clean)

    if all_data_frames:
        df_master = pd.concat(all_data_frames, ignore_index=True)

        st.subheader(f"✅ Berhasil Mengumpulkan {len(df_master)} Baris Data")
        st.dataframe(df_master.head(20))

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_master.to_excel(writer, index=False, sheet_name='Master_Data_Zizah')

        st.download_button(
            label="📥 Download Master Data Lengkap (Excel)",
            data=buffer.getvalue(),
            file_name='MASTER_DATA_VLOOKUP_ZIZAH.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        st.success("Download file di atas, isinya sudah daftar nama lengkap untuk bahan VLOOKUP. 🥂")
    else:
        st.error("⚠️ Tidak ditemukan kolom 'NAMA'. Pastikan header di Excel tulisannya 'NAMA'.")

# --- ZONA TESTING ---
st.markdown("---")
st.subheader("🛠️ Zona Testing (Gunakan ini kalau mau tes hasil)")
st.info("Download file di bawah, lalu upload kembali ke atas untuk tes.")

def buat_file_test_maret():
    wb = Workbook()
    ws = wb.active
    ws.title = "BA_Maret"
    ws['A1'] = 'PT KRAKATAU BAJA'
    ws['A2'] = 'BERITA ACARA KEHADIRAN TENAGA KERJA'
    ws['A3'] = 'BULAN: MARET 2026'
    ws['A4'] = 'LOKASI KERJA: CILEGON'
    ws['A5'] = ''
    for col, h in enumerate(['NO', 'NIK', 'NAMA', 'JABATAN', 'PERUSAHAAN'], 1):
        ws.cell(6, col, h)
    karyawan = [
        (1, 'A001', 'Zizah Nur Aini',  'Admin',      'PT Maju Jaya'),
        (2, 'A002', 'Budi Santoso',     'Teknisi',    'PT Maju Jaya'),
        (3, 'A003', 'Ani Rahayu',       'Logistik',   'PT Sejahtera'),
        (4, 'A004', 'Candra Wijaya',    'Supervisor', 'PT Sejahtera'),
        (5, 'A005', 'Dewi Lestari',     'Operator',   'PT Karya Utama'),
    ]
    for i, row in enumerate(karyawan, 7):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def buat_file_test_april():
    wb = Workbook()
    ws = wb.active
    ws.title = "BA_April"
    ws['A1'] = 'BERITA ACARA KEHADIRAN - APRIL 2026'
    ws['A2'] = 'KONTRAKTOR: PT MAJU BERSAMA'
    ws['A3'] = 'LOKASI: CILEGON'
    for col, h in enumerate(['NO', 'NIK', 'NAMA', 'JABATAN', 'PERUSAHAAN'], 1):
        ws.cell(4, col, h)
    karyawan = [
        (1, 'B001', 'Dedi Kurniawan',    'Driver',   'PT Logindo'),
        (2, 'B002', 'Eka Putri Mandiri', 'Security', 'PT Logindo'),
        (3, 'B003', 'Faisal Rahman',     'Helper',   'PT Konstruksi'),
        (4, 'B004', 'Gina Marlina',      'Operator', 'PT Konstruksi'),
    ]
    for i, row in enumerate(karyawan, 5):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

col1, col2 = st.columns(2)

with col1:
    st.markdown("**File Test A — Maret** (5 karyawan)")
    st.download_button(
        label="📥 Download File Test Maret",
        data=buat_file_test_maret(),
        file_name="Test_BA_Maret.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with col2:
    st.markdown("**File Test B — April** (4 karyawan)")
    st.download_button(
        label="📥 Download File Test April",
        data=buat_file_test_april(),
        file_name="Test_BA_April.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.caption("Setelah download kedua file, upload ke uploader di atas untuk melihat hasil rekap.")
